from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO

import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st
from matplotlib.colors import BoundaryNorm, ListedColormap, to_hex
from matplotlib.lines import Line2D
from openpyxl import load_workbook

DEFAULT_HEATMAP_COLORS = {
    "strong_down": "#2f8ac4",
    "mild_down": "#cfe8f6",
    "neutral": "#f1d783",
    "mild_up": "#f68b63",
    "strong_up": "#ff646b",
    "pvalue": "#eae9e9",
}

HEATMAP_COLOR_WIDGET_VERSION = "v5"

DEFAULT_CLASS_COLORS = [
    "#000eff",
    "#bdd9ff",
    "#ff7600",
    "#fdc68e",
    "#06a221",
    "#73ff56",
    "#ff0001",
    "#ffa29e",
    "#8500ff",
    "#d197ff",
    "#e4ff00",
    "#ffe300",
    "#ff00b3",
]

CLASS_COLOR_WIDGET_VERSION = "v2"

DEFAULT_METRIC_GROUP_COLORS = [
    "#98f3d5",
    "#ffa580",
    "#a4b5ea",
    "#f7b2d6",
    "#c0ea77",
    "#ffef8e",
    "#dc5578",
    "#9E91FF",
]

METRIC_GROUP_COLOR_WIDGET_VERSION = "v3"

DEFAULT_COMPOUND_TYPE_COLORS = [
    "#e35d6a",
    "#54b7c8",
    "#b8b8b8",
]

COMPOUND_TYPE_COLOR_WIDGET_VERSION = "v1"


@dataclass
class ParsedTable:
    data: pd.DataFrame
    metrics: list[str]
    metric_groups: dict[str, str]
    has_metric_groups: bool


def normalize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def read_merged_value(sheet, row: int, col: int) -> object:
    cell = sheet.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value

    for merged_range in sheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return None


@st.cache_data(show_spinner=False)
def parse_workbook(file_bytes: bytes) -> ParsedTable:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    subheader_row = 2
    for row in range(1, min(sheet.max_row, 6) + 1):
        fc_header = normalize_text(read_merged_value(sheet, row, 3))
        pvalue_header = normalize_text(read_merged_value(sheet, row, 4))
        if (
            fc_header
            and pvalue_header
            and fc_header.lower() == "log2fold change"
            and pvalue_header.lower() == "p-value"
        ):
            subheader_row = row
            break

    metric_row = subheader_row - 1
    metric_group_row = subheader_row - 2
    has_metric_groups = metric_group_row >= 1 and any(
        normalize_text(read_merged_value(sheet, metric_group_row, col))
        for col in range(3, sheet.max_column + 1, 2)
    )

    metric_columns: list[tuple[str, int, int, str]] = []
    metric_groups: dict[str, str] = {}
    for col in range(3, sheet.max_column + 1, 2):
        metric_name = normalize_text(read_merged_value(sheet, metric_row, col))
        if not metric_name:
            continue
        metric_group = ""
        if has_metric_groups:
            metric_group = (
                normalize_text(read_merged_value(sheet, metric_group_row, col))
                or "Ungrouped"
            )
        metric_groups[metric_name] = metric_group
        metric_columns.append((metric_name, col, col + 1, metric_group))

    rows: list[dict[str, object]] = []
    current_group: str | None = None

    for row in range(subheader_row + 1, sheet.max_row + 1):
        group = normalize_text(read_merged_value(sheet, row, 1))
        drug = normalize_text(sheet.cell(row=row, column=2).value)

        if group:
            current_group = group
        if not drug:
            continue

        record: dict[str, object] = {
            "group": current_group or "Uncategorized",
            "drug": drug,
        }

        has_any_value = False
        for metric_name, fc_col, p_col, _ in metric_columns:
            fc_value = sheet.cell(row=row, column=fc_col).value
            p_value = sheet.cell(row=row, column=p_col).value

            fc_numeric = pd.to_numeric(fc_value, errors="coerce")
            p_numeric = pd.to_numeric(p_value, errors="coerce")

            record[f"{metric_name}__log2fc"] = fc_numeric
            record[f"{metric_name}__pvalue"] = p_numeric
            has_any_value = has_any_value or pd.notna(fc_numeric)

        if has_any_value:
            rows.append(record)

    data = pd.DataFrame(rows)
    workbook.close()
    return ParsedTable(
        data=data,
        metrics=[name for name, _, _, _ in metric_columns],
        metric_groups=metric_groups,
        has_metric_groups=has_metric_groups,
    )


@st.cache_data(show_spinner=False)
def parse_compound_workbook(file_bytes: bytes) -> ParsedTable:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    subheader_row = 3
    first_metric_col = 4
    for row in range(1, min(sheet.max_row, 8) + 1):
        for col in range(1, sheet.max_column):
            fc_header = normalize_text(read_merged_value(sheet, row, col))
            pvalue_header = normalize_text(read_merged_value(sheet, row, col + 1))
            if (
                fc_header
                and pvalue_header
                and fc_header.lower() == "log2fold change"
                and pvalue_header.lower() == "p-value"
            ):
                subheader_row = row
                first_metric_col = col
                break
        else:
            continue
        break

    metric_row = subheader_row - 1
    metric_group_row = subheader_row - 2
    has_metric_groups = metric_group_row >= 1 and any(
        normalize_text(read_merged_value(sheet, metric_group_row, col))
        for col in range(first_metric_col, sheet.max_column + 1, 2)
    )

    metric_columns: list[tuple[str, int, int, str]] = []
    metric_groups: dict[str, str] = {}
    for col in range(first_metric_col, sheet.max_column + 1, 2):
        metric_name = normalize_text(read_merged_value(sheet, metric_row, col))
        if not metric_name:
            continue
        metric_group = ""
        if has_metric_groups:
            metric_group = (
                normalize_text(read_merged_value(sheet, metric_group_row, col))
                or "Ungrouped"
            )
        metric_groups[metric_name] = metric_group
        metric_columns.append((metric_name, col, col + 1, metric_group))

    rows: list[dict[str, object]] = []
    current_compound_type: str | None = None
    current_group: str | None = None

    for row in range(subheader_row + 1, sheet.max_row + 1):
        compound_type = normalize_text(read_merged_value(sheet, row, 1))
        group = normalize_text(read_merged_value(sheet, row, 2))
        drug = normalize_text(sheet.cell(row=row, column=3).value)

        if compound_type:
            current_compound_type = compound_type
        if group:
            current_group = group
        if not drug:
            continue

        record: dict[str, object] = {
            "compound_type": current_compound_type or "Uncategorized",
            "group": current_group or "Uncategorized",
            "drug": drug,
        }

        has_any_value = False
        for metric_name, fc_col, p_col, _ in metric_columns:
            fc_value = sheet.cell(row=row, column=fc_col).value
            p_value = sheet.cell(row=row, column=p_col).value

            fc_numeric = pd.to_numeric(fc_value, errors="coerce")
            p_numeric = pd.to_numeric(p_value, errors="coerce")

            record[f"{metric_name}__log2fc"] = fc_numeric
            record[f"{metric_name}__pvalue"] = p_numeric
            has_any_value = has_any_value or pd.notna(fc_numeric)

        if has_any_value:
            rows.append(record)

    data = pd.DataFrame(rows)
    workbook.close()
    return ParsedTable(
        data=data,
        metrics=[name for name, _, _, _ in metric_columns],
        metric_groups=metric_groups,
        has_metric_groups=has_metric_groups,
    )


def build_heatmap_frame(
    data: pd.DataFrame,
    selected_groups: list[str],
    selected_metrics: list[str],
    pvalue_threshold: float,
    significant_only: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    filtered = data[data["group"].isin(selected_groups)].reset_index(drop=True).copy()

    if filtered.empty:
        return filtered, pd.DataFrame(), pd.DataFrame()

    value_frame = pd.DataFrame(
        {metric: filtered[f"{metric}__log2fc"].to_numpy() for metric in selected_metrics}
    )
    pvalue_frame = pd.DataFrame(
        {metric: filtered[f"{metric}__pvalue"].to_numpy() for metric in selected_metrics}
    )

    if significant_only:
        significant_mask = (pvalue_frame <= pvalue_threshold).fillna(False).any(axis=1)
        filtered = filtered.loc[significant_mask].reset_index(drop=True)
        value_frame = value_frame.loc[significant_mask].reset_index(drop=True)
        pvalue_frame = pvalue_frame.loc[significant_mask].reset_index(drop=True)

    return filtered, value_frame, pvalue_frame


def build_compound_heatmap_frame(
    data: pd.DataFrame,
    selected_compound_types: list[str],
    selected_groups: list[str],
    selected_metrics: list[str],
    pvalue_threshold: float,
    significant_only: bool,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    filtered = data[
        data["compound_type"].isin(selected_compound_types)
        & data["group"].isin(selected_groups)
    ].reset_index(drop=True).copy()

    if filtered.empty:
        return filtered, pd.DataFrame(), pd.DataFrame()

    value_frame = pd.DataFrame(
        {metric: filtered[f"{metric}__log2fc"].to_numpy() for metric in selected_metrics}
    )
    pvalue_frame = pd.DataFrame(
        {metric: filtered[f"{metric}__pvalue"].to_numpy() for metric in selected_metrics}
    )

    if significant_only:
        significant_mask = (pvalue_frame <= pvalue_threshold).fillna(False).any(axis=1)
        filtered = filtered.loc[significant_mask].reset_index(drop=True)
        value_frame = value_frame.loc[significant_mask].reset_index(drop=True)
        pvalue_frame = pvalue_frame.loc[significant_mask].reset_index(drop=True)

    return filtered, value_frame, pvalue_frame


def editable_text_input(label: str, key: str, default: str) -> str:
    if key not in st.session_state:
        st.session_state[key] = default
    return st.text_input(label, key=key)


def build_default_text_labels(fc_cutoff: float, pvalue_threshold: float) -> dict[str, str]:
    return {
        "heatmap_title": "",
        "class_strip_title": "Drug classes",
        "legend_title": "Drug classes",
        "metric_group_strip_title": "Functional groups",
        "metric_group_legend_title": "Functional groups",
        "colorbar_title": "Log2 Fold Change",
        "pvalue_note": f"p-value > {pvalue_threshold:g}",
        "bin_strong_down": f"<= -{fc_cutoff:g}",
        "bin_mild_down": f"-{fc_cutoff:g} — -0.5",
        "bin_neutral": "-0.5 — 0.5",
        "bin_mild_up": f"0.5 — {fc_cutoff:g}",
        "bin_strong_up": f">= {fc_cutoff:g}",
    }


def sync_default_text_labels(
    default_text_labels: dict[str, str],
    state_prefix: str = "text_",
    tracker_key: str = "text_default_tracker",
) -> None:
    previous_defaults = st.session_state.get(tracker_key, {})

    for label_key, current_default in default_text_labels.items():
        state_key = f"{state_prefix}{label_key}"
        previous_default = previous_defaults.get(label_key)
        if state_key not in st.session_state or st.session_state[state_key] == previous_default:
            st.session_state[state_key] = current_default

    st.session_state[tracker_key] = default_text_labels.copy()


def editable_mapping_editor(
    items: list[str],
    state_key: str,
    source_label: str,
    target_label: str,
) -> dict[str, str]:
    if state_key not in st.session_state:
        st.session_state[state_key] = {}

    mapping = st.session_state[state_key]
    for item in items:
        mapping.setdefault(item, item)

    editor_df = pd.DataFrame(
        {
            source_label: items,
            target_label: [mapping[item] for item in items],
        }
    )
    edited_df = st.data_editor(
        editor_df,
        hide_index=True,
        width="stretch",
        num_rows="fixed",
        key=f"{state_key}_editor",
        column_config={
            source_label: st.column_config.TextColumn(disabled=True),
            target_label: st.column_config.TextColumn(required=True),
        },
    )

    updated_mapping = {
        row[source_label]: normalize_text(row[target_label]) or row[source_label]
        for _, row in edited_df.iterrows()
    }
    st.session_state[state_key].update(updated_mapping)
    return st.session_state[state_key]


def build_palette_color_map(items: list[str], palette_name: str) -> dict[str, str]:
    palette = plt.get_cmap(palette_name)
    return {
        item: to_hex(palette(index % palette.N))
        for index, item in enumerate(items)
    }


def build_ordered_color_map(
    items: list[str],
    colors: list[str],
    fallback_palette_name: str,
) -> dict[str, str]:
    fallback_palette = plt.get_cmap(fallback_palette_name)
    return {
        item: (
            colors[index]
            if index < len(colors)
            else to_hex(fallback_palette(index % fallback_palette.N))
        )
        for index, item in enumerate(items)
    }


def editable_color_map(
    items: list[str],
    default_colors: dict[str, str],
    key_prefix: str,
    label_map: dict[str, str] | None = None,
) -> dict[str, str]:
    colors: dict[str, str] = {}
    for item in items:
        display_label = label_map.get(item, item) if label_map else item
        colors[item] = st.color_picker(
            str(display_label),
            value=default_colors.get(item, "#cccccc"),
            key=f"{key_prefix}_{item}",
        )
    return colors


def plot_heatmap(
    values: pd.DataFrame,
    pvalues: pd.DataFrame,
    pvalue_threshold: float,
    low_cutoff: float,
    high_cutoff: float,
    annotate: bool,
    text_labels: dict[str, str],
    metric_labels: list[str],
    metric_group_labels: list[str],
    drug_labels: list[str],
    group_labels: list[str],
    show_metric_groups: bool,
    heatmap_colors: list[str],
    pvalue_color: str,
    class_color_map: dict[str, str],
    metric_group_color_map: dict[str, str],
    orientation: str,
) -> plt.Figure:
    shown_values = values.copy()
    shown_values = shown_values.where(pvalues <= pvalue_threshold, np.nan)

    color_steps = heatmap_colors
    cmap = ListedColormap(color_steps)
    finite_values = values.to_numpy(dtype=float)
    finite_values = finite_values[np.isfinite(finite_values)]
    max_abs_value = float(np.max(np.abs(finite_values))) if finite_values.size else 2.0
    outer_limit = max(max_abs_value, abs(low_cutoff), abs(high_cutoff), 2.0) + 0.01
    bounds = [-outer_limit, low_cutoff, -0.5, 0.5, high_cutoff, outer_limit]
    norm = BoundaryNorm(bounds, cmap.N, clip=True)

    unique_groups = list(dict.fromkeys(group_labels))
    palette = plt.get_cmap("tab20")
    group_to_color = {
        group: class_color_map.get(group, to_hex(palette(i % palette.N)))
        for i, group in enumerate(unique_groups)
    }
    group_colors = [group_to_color[group] for group in group_labels]
    unique_metric_groups = list(dict.fromkeys(metric_group_labels)) if show_metric_groups else []
    metric_palette = plt.get_cmap("Set2")
    metric_group_to_color = {
        group: metric_group_color_map.get(group, to_hex(metric_palette(i % metric_palette.N)))
        for i, group in enumerate(unique_metric_groups)
    }
    metric_group_colors = (
        [metric_group_to_color[group] for group in metric_group_labels]
        if show_metric_groups
        else []
    )

    is_transposed = orientation == "groups_top"
    if is_transposed:
        plot_values = shown_values.T
        raw_values = values.T
        plot_pvalues = pvalues.T
        x_labels = drug_labels
        y_labels = metric_labels
        max_drug_label_len = max((len(label) for label in drug_labels), default=1)
        label_space = max(0.75, max_drug_label_len * 0.065 + 0.35)
        strip_space = 0.22
        top_space = 0.28
        legend_space = 1.7 if show_metric_groups else 1.35
        fig_height = max(
            5.5,
            0.45 * len(metric_labels) + 2.1 + max(0.0, label_space - 0.9),
        ) + legend_space
        fig_width = max(11.5, 0.45 * len(drug_labels) + 3.2)
        fig = plt.figure(figsize=(fig_width, fig_height))
        legend_bottom = 0.035
        legend_height = legend_space / fig_height
        heatmap_bottom = legend_bottom + legend_height + 0.04
        heatmap_left = 0.22
        heatmap_width = 0.72
        metric_group_width = 0.018 if show_metric_groups else 0.0
        metric_group_gap = 0.006 if show_metric_groups else 0.0
        heatmap_top = 1 - (top_space + strip_space + label_space) / fig_height
        heatmap_height = heatmap_top - heatmap_bottom

        ax_metric_group = None
        ax_heatmap = fig.add_axes(
            [heatmap_left, heatmap_bottom, heatmap_width, heatmap_height]
        )
        ax_legend = fig.add_axes(
            [heatmap_left, legend_bottom, heatmap_width, legend_height]
        )
        ax_group = None
        group_matrix = np.arange(len(group_labels)).reshape(1, -1)
        metric_group_matrix = np.arange(len(metric_group_labels)).reshape(-1, 1)
    else:
        plot_values = shown_values
        raw_values = values
        plot_pvalues = pvalues
        x_labels = metric_labels
        y_labels = drug_labels
        n_rows = max(len(values), 1)
        n_cols = max(len(values.columns), 1)
        max_metric_label_len = max((len(label) for label in metric_labels), default=1)
        label_space = max(0.75, max_metric_label_len * 0.065 + 0.35)
        strip_space = 0.22
        top_space = 0.22
        legend_space = 1.7 if show_metric_groups else 1.35
        fig_height = max(7.0, 0.48 * n_rows + 3.1) + legend_space
        fig_width = max(10.5, 0.46 * n_cols + 3.0)
        fig = plt.figure(figsize=(fig_width, fig_height))
        legend_bottom = 0.035
        legend_height = legend_space / fig_height
        legend_gap = 0.45 / fig_height
        heatmap_bottom = legend_bottom + legend_height + legend_gap
        heatmap_left = 0.22
        heatmap_width = 0.73
        heatmap_top = 1 - (top_space + strip_space + label_space) / fig_height
        heatmap_height = heatmap_top - heatmap_bottom

        ax_metric_group = None
        ax_group = None
        ax_heatmap = fig.add_axes(
            [heatmap_left, heatmap_bottom, heatmap_width, heatmap_height]
        )
        ax_legend = fig.add_axes(
            [heatmap_left, legend_bottom, heatmap_width, legend_height]
        )
        group_matrix = np.arange(len(group_labels)).reshape(-1, 1)
        metric_group_matrix = np.arange(len(metric_group_labels)).reshape(1, -1)

    masked = np.ma.masked_invalid(plot_values.to_numpy(dtype=float))
    cmap_with_nan = cmap.with_extremes(bad=pvalue_color)
    image = ax_heatmap.imshow(masked, aspect="auto", cmap=cmap_with_nan, norm=norm)

    ax_heatmap.set_xticks(np.arange(len(x_labels)))
    if is_transposed:
        ax_heatmap.xaxis.tick_top()
        ax_heatmap.set_xticklabels(
            x_labels,
            rotation=45,
            ha="left",
            rotation_mode="anchor",
            fontsize=10,
            fontweight="bold",
        )
        ax_heatmap.tick_params(axis="x", labeltop=True, labelbottom=False, pad=2)
    else:
        ax_heatmap.set_xticklabels(
            x_labels,
            rotation=45,
            ha="left",
            rotation_mode="anchor",
            fontsize=10,
            fontweight="bold",
        )
        ax_heatmap.xaxis.tick_top()
        ax_heatmap.tick_params(axis="x", labeltop=True, labelbottom=False, pad=2)
    ax_heatmap.set_yticks(np.arange(len(y_labels)))
    ax_heatmap.set_yticklabels(y_labels, fontsize=10, fontweight="bold")
    ax_heatmap.tick_params(length=0)
    if text_labels["heatmap_title"]:
        ax_heatmap.set_title(text_labels["heatmap_title"], loc="left", fontsize=16, pad=12)

    ax_heatmap.set_xticks(np.arange(-0.5, len(x_labels), 1), minor=True)
    ax_heatmap.set_yticks(np.arange(-0.5, len(y_labels), 1), minor=True)
    ax_heatmap.grid(which="major", visible=False)
    ax_heatmap.grid(which="minor", color="black", linestyle="-", linewidth=1)
    for spine in ax_heatmap.spines.values():
        spine.set_visible(False)

    group_cmap = ListedColormap(group_colors or ["#cccccc"])
    metric_group_cmap = ListedColormap(metric_group_colors or ["#cccccc"])
    if is_transposed:
        if show_metric_groups:
            fig.canvas.draw()
            renderer = fig.canvas.get_renderer()
            tick_label_lefts = [
                label.get_window_extent(renderer=renderer).transformed(
                    fig.transFigure.inverted()
                ).x0
                for label in ax_heatmap.get_yticklabels()
                if label.get_visible()
            ]
            label_left = min(tick_label_lefts, default=ax_heatmap.get_position().x0)
            metric_group_left = label_left - metric_group_gap - metric_group_width
            min_metric_group_left = 0.055
            if metric_group_left < min_metric_group_left:
                heatmap_position = ax_heatmap.get_position()
                shift = min_metric_group_left - metric_group_left
                new_heatmap_left = heatmap_position.x0 + shift
                new_heatmap_width = max(0.35, heatmap_position.width - shift)
                ax_heatmap.set_position(
                    [
                        new_heatmap_left,
                        heatmap_position.y0,
                        new_heatmap_width,
                        heatmap_position.height,
                    ]
                )
                heatmap_left = new_heatmap_left
                heatmap_width = new_heatmap_width

        strip_height = strip_space / fig_height
        strip_gap = 0.08 / fig_height
        top_margin = 0.22 / fig_height
        max_strip_bottom = 1 - strip_height - top_margin

        def get_top_tick_label_edge() -> float:
            fig.canvas.draw()
            renderer = fig.canvas.get_renderer()
            tick_label_tops = [
                label.get_window_extent(renderer=renderer).transformed(
                    fig.transFigure.inverted()
                ).y1
                for label in ax_heatmap.get_xticklabels()
                if label.get_visible()
            ]
            return max(tick_label_tops, default=ax_heatmap.get_position().y1)

        label_top = get_top_tick_label_edge()
        strip_bottom = label_top + strip_gap
        if strip_bottom > max_strip_bottom:
            heatmap_position = ax_heatmap.get_position()
            overflow = strip_bottom - max_strip_bottom
            new_height = max(0.25, heatmap_position.height - overflow)
            ax_heatmap.set_position(
                [
                    heatmap_position.x0,
                    heatmap_position.y0,
                    heatmap_position.width,
                    new_height,
                ]
            )
            if ax_metric_group is not None:
                metric_group_position = ax_metric_group.get_position()
                ax_metric_group.set_position(
                    [
                        metric_group_position.x0,
                        metric_group_position.y0,
                        metric_group_position.width,
                        new_height,
                    ]
                )
            label_top = get_top_tick_label_edge()
            strip_bottom = label_top + strip_gap

        strip_bottom = min(strip_bottom, max_strip_bottom)
        ax_group = fig.add_axes(
            [heatmap_left, strip_bottom, heatmap_width, strip_height]
        )
        ax_legend.set_position(
            [heatmap_left, legend_bottom, heatmap_width, legend_height]
        )
        if show_metric_groups:
            fig.canvas.draw()
            renderer = fig.canvas.get_renderer()
            tick_label_lefts = [
                label.get_window_extent(renderer=renderer).transformed(
                    fig.transFigure.inverted()
                ).x0
                for label in ax_heatmap.get_yticklabels()
                if label.get_visible()
            ]
            label_left = min(tick_label_lefts, default=ax_heatmap.get_position().x0)
            metric_group_right = label_left - metric_group_gap
            metric_group_left = metric_group_right - metric_group_width
            heatmap_position = ax_heatmap.get_position()
            ax_metric_group = fig.add_axes(
                [
                    metric_group_left,
                    heatmap_position.y0,
                    metric_group_width,
                    heatmap_position.height,
                ]
            )
    else:
        strip_height = strip_space / fig_height
        strip_gap = 0.025 / fig_height
        side_strip_width = strip_space / fig_width
        side_strip_gap = 0.025 / fig_width

        def get_y_tick_label_left() -> float:
            fig.canvas.draw()
            renderer = fig.canvas.get_renderer()
            tick_label_lefts = [
                label.get_window_extent(renderer=renderer).transformed(
                    fig.transFigure.inverted()
                ).x0
                for label in ax_heatmap.get_yticklabels()
                if label.get_visible()
            ]
            return min(tick_label_lefts, default=ax_heatmap.get_position().x0)

        label_left = get_y_tick_label_left()
        side_strip_left = label_left - side_strip_gap - side_strip_width
        min_side_strip_left = 0.035
        if side_strip_left < min_side_strip_left:
            heatmap_position = ax_heatmap.get_position()
            shift = min_side_strip_left - side_strip_left
            ax_heatmap.set_position(
                [
                    heatmap_position.x0 + shift,
                    heatmap_position.y0,
                    max(0.35, heatmap_position.width - shift),
                    heatmap_position.height,
                ]
            )
            label_left = get_y_tick_label_left()

        heatmap_position = ax_heatmap.get_position()
        side_strip_right = label_left - side_strip_gap
        side_strip_left = side_strip_right - side_strip_width
        ax_group = fig.add_axes(
            [
                side_strip_left,
                heatmap_position.y0,
                side_strip_width,
                heatmap_position.height,
            ]
        )

        if show_metric_groups:
            top_margin = 0.18 / fig_height
            max_strip_bottom = 1 - strip_height - top_margin

            def get_top_tick_label_edge() -> float:
                fig.canvas.draw()
                renderer = fig.canvas.get_renderer()
                tick_label_tops = [
                    label.get_window_extent(renderer=renderer).transformed(
                        fig.transFigure.inverted()
                    ).y1
                    for label in ax_heatmap.get_xticklabels()
                    if label.get_visible()
                ]
                return max(tick_label_tops, default=ax_heatmap.get_position().y1)

            label_top = get_top_tick_label_edge()
            strip_bottom = label_top + strip_gap
            if strip_bottom > max_strip_bottom:
                heatmap_position = ax_heatmap.get_position()
                overflow = strip_bottom - max_strip_bottom
                new_height = max(0.25, heatmap_position.height - overflow)
                ax_heatmap.set_position(
                    [
                        heatmap_position.x0,
                        heatmap_position.y0,
                        heatmap_position.width,
                        new_height,
                    ]
                )
                group_position = ax_group.get_position()
                ax_group.set_position(
                    [
                        group_position.x0,
                        group_position.y0,
                        group_position.width,
                        new_height,
                    ]
                )
                label_top = get_top_tick_label_edge()
                strip_bottom = label_top + strip_gap

            strip_bottom = min(strip_bottom, max_strip_bottom)
            heatmap_position = ax_heatmap.get_position()
            ax_metric_group = fig.add_axes(
                [
                    heatmap_position.x0,
                    strip_bottom,
                    heatmap_position.width,
                    strip_height,
                ]
            )

        heatmap_position = ax_heatmap.get_position()
        ax_legend.set_position(
            [heatmap_position.x0, legend_bottom, heatmap_position.width, legend_height]
        )

    ax_group.imshow(group_matrix, aspect="auto", cmap=group_cmap)
    ax_group.set_xticks([])
    ax_group.set_yticks([])
    ax_group.tick_params(left=False, bottom=False, labelleft=False)
    if is_transposed:
        ax_group.set_title(
            text_labels["class_strip_title"],
            fontsize=10,
            pad=16,
            fontweight="bold",
        )
    else:
        ax_group.set_ylabel(
            text_labels["class_strip_title"],
            fontsize=10,
            rotation=90,
            labelpad=8,
            va="center",
            fontweight="bold",
        )
    for spine in ax_group.spines.values():
        spine.set_visible(False)

    if show_metric_groups and ax_metric_group is not None:
        ax_metric_group.imshow(
            metric_group_matrix,
            aspect="auto",
            cmap=metric_group_cmap,
        )
        if is_transposed:
            ax_metric_group.set_ylim(ax_heatmap.get_ylim())
            ax_metric_group.set_xticks([])
            ax_metric_group.set_yticks([])
            ax_metric_group.tick_params(axis="y", left=False, labelleft=False, length=0)
            ax_metric_group.tick_params(axis="x", bottom=False, labelbottom=False)
            ax_metric_group.set_ylabel(
                text_labels["metric_group_strip_title"],
                fontsize=10,
                rotation=90,
                labelpad=16,
                va="center",
                fontweight="bold",
            )
        else:
            ax_metric_group.set_xlim(ax_heatmap.get_xlim())
            ax_metric_group.set_xticks([])
            ax_metric_group.set_yticks([])
            ax_metric_group.tick_params(
                left=False,
                bottom=False,
                labelleft=False,
                labelbottom=False,
            )
            ax_metric_group.set_title(
                text_labels["metric_group_strip_title"],
                fontsize=10,
                pad=8,
                fontweight="bold",
            )
        for spine in ax_metric_group.spines.values():
            spine.set_visible(False)

    if annotate:
        for row_idx in range(len(plot_values.index)):
            for col_idx in range(len(plot_values.columns)):
                raw_value = raw_values.iloc[row_idx, col_idx]
                pval = plot_pvalues.iloc[row_idx, col_idx]
                if pd.isna(raw_value):
                    continue
                if pd.isna(pval) or pval > pvalue_threshold:
                    continue

                ax_heatmap.text(
                    col_idx,
                    row_idx,
                    f"{raw_value:.2f}",
                    ha="center",
                    va="center",
                    fontsize=8.5 if not is_transposed else 7.5,
                    fontweight="bold",
                    color="black",
                )

    fc_handles = [
        mpatches.Patch(
            facecolor=color_steps[0],
            edgecolor="black",
            linewidth=0.5,
            label=text_labels["bin_strong_down"],
        ),
        mpatches.Patch(
            facecolor=color_steps[1],
            edgecolor="black",
            linewidth=0.5,
            label=text_labels["bin_mild_down"],
        ),
        mpatches.Patch(
            facecolor=color_steps[2],
            edgecolor="black",
            linewidth=0.5,
            label=text_labels["bin_neutral"],
        ),
        mpatches.Patch(
            facecolor=color_steps[3],
            edgecolor="black",
            linewidth=0.5,
            label=text_labels["bin_mild_up"],
        ),
        mpatches.Patch(
            facecolor=color_steps[4],
            edgecolor="black",
            linewidth=0.5,
            label=text_labels["bin_strong_up"],
        ),
        mpatches.Patch(
            facecolor=pvalue_color,
            edgecolor="black",
            linewidth=0.5,
            label=text_labels["pvalue_note"],
        ),
    ]
    class_handles = [
        mpatches.Patch(
            facecolor=color,
            edgecolor="black",
            linewidth=0.5,
            label=group,
        )
        for group, color in group_to_color.items()
    ]
    metric_group_handles = [
        mpatches.Patch(
            facecolor=color,
            edgecolor="black",
            linewidth=0.5,
            label=group,
        )
        for group, color in metric_group_to_color.items()
    ]

    fig.canvas.draw()
    heatmap_position = ax_heatmap.get_position()
    legend_position = ax_legend.get_position()
    ax_legend.set_position(
        [
            heatmap_position.x0,
            legend_position.y0,
            heatmap_position.width,
            legend_position.height,
        ]
    )
    ax_legend.axis("off")

    def place_legend(
        handles: list[mpatches.Patch],
        title: str,
        anchor_x: float,
        columns: int = 1,
    ) -> plt.Legend:
        legend = ax_legend.legend(
            handles=handles,
            loc="upper left",
            bbox_to_anchor=(anchor_x, 1.0),
            borderaxespad=0,
            borderpad=0,
            title=title,
            frameon=False,
            fontsize=9,
            title_fontsize=10,
            ncol=columns,
            columnspacing=1.0,
            handlelength=1.7,
            handletextpad=0.55,
            labelspacing=0.35,
        )
        legend._legend_box.align = "left"
        legend.get_title().set_ha("left")
        return legend

    fc_legend = place_legend(fc_handles, text_labels["colorbar_title"], 0.0)
    ax_legend.add_artist(fc_legend)

    class_columns = 2 if len(class_handles) > 6 else 1
    class_anchor = 0.24 if show_metric_groups else 0.36
    class_legend = place_legend(
        class_handles,
        text_labels["legend_title"],
        class_anchor,
        columns=class_columns,
    )
    if show_metric_groups:
        ax_legend.add_artist(class_legend)
        metric_group_legend = place_legend(
            metric_group_handles,
            text_labels["metric_group_legend_title"],
            0.62,
        )

    return fig


def plot_compound_dot_heatmap(
    values: pd.DataFrame,
    pvalues: pd.DataFrame,
    pvalue_threshold: float,
    low_cutoff: float,
    high_cutoff: float,
    text_labels: dict[str, str],
    metric_labels: list[str],
    metric_group_labels: list[str],
    drug_labels: list[str],
    group_labels: list[str],
    compound_type_labels: list[str],
    show_metric_groups: bool,
    heatmap_colors: list[str],
    pvalue_color: str,
    class_color_map: dict[str, str],
    metric_group_color_map: dict[str, str],
) -> plt.Figure:
    color_steps = heatmap_colors
    plot_values = values.T
    plot_pvalues = pvalues.T
    n_rows = max(len(plot_values.index), 1)
    n_cols = max(len(plot_values.columns), 1)

    compound_segments: list[tuple[int, int, str]] = []
    start = 0
    for index in range(1, len(compound_type_labels) + 1):
        if (
            index == len(compound_type_labels)
            or compound_type_labels[index] != compound_type_labels[start]
        ):
            compound_segments.append((start, index, compound_type_labels[start]))
            start = index

    gap_cells = 1.25
    x_positions: list[float] = []
    x_offset = 0.0
    for index in range(n_cols):
        if index > 0 and compound_type_labels[index] != compound_type_labels[index - 1]:
            x_offset += gap_cells
        x_positions.append(index + x_offset)
    plot_width_cells = n_cols + gap_cells * max(0, len(compound_segments) - 1)
    x_min = -0.5
    x_max = (x_positions[-1] + 0.5) if x_positions else 0.5

    cell_size = 0.42
    legend_space = 2.55
    legend_bottom_space = 0.35
    legend_gap = 0.75
    max_metric_label_len = max((len(label) for label in metric_labels), default=1)
    max_drug_label_len = max((len(label) for label in drug_labels), default=1)
    top_label_space = max(1.95, max_drug_label_len * 0.07 + 1.55)
    left_label_space = max(2.2, max_metric_label_len * 0.09 + 0.9)
    fig_width = max(11.0, left_label_space + plot_width_cells * cell_size + 0.9)
    fig_height = max(
        7.0,
        top_label_space + n_rows * cell_size + legend_space + legend_gap + legend_bottom_space,
    )

    fig = plt.figure(figsize=(fig_width, fig_height))
    heatmap_left = left_label_space / fig_width
    heatmap_bottom = (legend_bottom_space + legend_space + legend_gap) / fig_height
    heatmap_width = (plot_width_cells * cell_size) / fig_width
    heatmap_height = (n_rows * cell_size) / fig_height
    ax_heatmap = fig.add_axes(
        [heatmap_left, heatmap_bottom, heatmap_width, heatmap_height]
    )
    ax_legend = fig.add_axes(
        [
            heatmap_left,
            legend_bottom_space / fig_height,
            heatmap_width,
            legend_space / fig_height,
        ]
    )

    ax_heatmap.set_xlim(x_min, x_max)
    ax_heatmap.set_ylim(n_rows - 0.5, -0.5)
    ax_heatmap.set_aspect("equal", adjustable="box")
    ax_heatmap.set_facecolor("#f8f8f8")

    ax_heatmap.set_xticks(x_positions)
    ax_heatmap.set_xticklabels(
        drug_labels,
        rotation=45,
        ha="left",
        rotation_mode="anchor",
        fontsize=10,
        fontweight="bold",
    )
    ax_heatmap.xaxis.tick_top()
    ax_heatmap.tick_params(axis="x", labeltop=True, labelbottom=False, pad=2)
    ax_heatmap.set_yticks(np.arange(n_rows))
    ax_heatmap.set_yticklabels(metric_labels, fontsize=10, fontweight="bold")
    ax_heatmap.tick_params(length=0)
    if text_labels["heatmap_title"]:
        ax_heatmap.set_title(text_labels["heatmap_title"], loc="left", fontsize=16, pad=12)

    ax_heatmap.grid(which="major", visible=False)
    for spine in ax_heatmap.spines.values():
        spine.set_visible(False)

    for row_idx in range(n_rows):
        for col_idx, x_position in enumerate(x_positions):
            ax_heatmap.add_patch(
                mpatches.Rectangle(
                    (x_position - 0.5, row_idx - 0.5),
                    1.0,
                    1.0,
                    facecolor="#f8f8f8",
                    edgecolor="black",
                    linewidth=0.85,
                    zorder=1,
                )
            )

    size_cutoff = max(abs(low_cutoff), abs(high_cutoff))
    dot_sizes = {
        "small": 42,
        "medium": 95,
        "large": 165,
    }

    def dot_size(value: float) -> int:
        abs_value = abs(value)
        if abs_value >= size_cutoff:
            return dot_sizes["large"]
        if abs_value >= 0.5:
            return dot_sizes["medium"]
        return dot_sizes["small"]

    def dot_color(value: float, pvalue: float) -> str:
        if pd.isna(pvalue) or pvalue > pvalue_threshold:
            return pvalue_color
        if value <= low_cutoff:
            return color_steps[0]
        if value < -0.5:
            return color_steps[1]
        if value <= 0.5:
            return color_steps[2]
        if value < high_cutoff:
            return color_steps[3]
        return color_steps[4]

    x_coords: list[float] = []
    y_coords: list[int] = []
    marker_sizes: list[int] = []
    marker_colors: list[object] = []
    marker_edges: list[object] = []
    marker_linewidths: list[float] = []
    for row_idx in range(len(plot_values.index)):
        for col_idx in range(len(plot_values.columns)):
            raw_value = plot_values.iloc[row_idx, col_idx]
            if pd.isna(raw_value):
                continue
            pvalue = plot_pvalues.iloc[row_idx, col_idx]
            numeric_value = float(raw_value)
            is_insignificant = pd.isna(pvalue) or pvalue > pvalue_threshold
            if is_insignificant:
                continue
            x_coords.append(x_positions[col_idx])
            y_coords.append(row_idx)
            marker_sizes.append(dot_size(numeric_value))
            marker_colors.append(dot_color(numeric_value, pvalue))
            marker_edges.append("black")
            marker_linewidths.append(0.45)

    ax_heatmap.scatter(
        x_coords,
        y_coords,
        s=marker_sizes,
        c=marker_colors,
        marker="o",
        edgecolors=marker_edges,
        linewidths=marker_linewidths,
        zorder=3,
    )

    unique_groups = list(dict.fromkeys(group_labels))
    group_palette = plt.get_cmap("tab20")
    group_to_color = {
        group: class_color_map.get(group, to_hex(group_palette(i % group_palette.N)))
        for i, group in enumerate(unique_groups)
    }
    group_colors = [group_to_color[group] for group in group_labels]

    unique_metric_groups = list(dict.fromkeys(metric_group_labels)) if show_metric_groups else []
    metric_palette = plt.get_cmap("Set2")
    metric_group_to_color = {
        group: metric_group_color_map.get(group, to_hex(metric_palette(i % metric_palette.N)))
        for i, group in enumerate(unique_metric_groups)
    }
    metric_group_colors = (
        [metric_group_to_color[group] for group in metric_group_labels]
        if show_metric_groups
        else []
    )

    group_cmap = ListedColormap(group_colors or ["#cccccc"])
    metric_group_cmap = ListedColormap(metric_group_colors or ["#cccccc"])

    fig.canvas.draw()
    strip_width = 0.18 / fig_width
    strip_height = 0.18 / fig_height
    strip_gap_x = 0.025 / fig_width
    strip_gap_y = 0.12 / fig_height

    def compound_type_color(label: str) -> str:
        normalized = label.lower()
        if "neurotoxic" in normalized:
            return "#d62728"
        if "neuroactive" in normalized:
            return "#1f77b4"
        if "reference" in normalized:
            return "#111111"
        return "#111111"

    def get_y_tick_label_left() -> float:
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        tick_label_lefts = [
            label.get_window_extent(renderer=renderer).transformed(
                fig.transFigure.inverted()
            ).x0
            for label in ax_heatmap.get_yticklabels()
            if label.get_visible()
        ]
        return min(tick_label_lefts, default=ax_heatmap.get_position().x0)

    label_left = get_y_tick_label_left()
    metric_group_right = label_left - strip_gap_x
    metric_group_left = metric_group_right - strip_width
    min_strip_left = 0.035
    if metric_group_left < min_strip_left:
        heatmap_position = ax_heatmap.get_position()
        shift = min_strip_left - metric_group_left
        ax_heatmap.set_position(
            [
                heatmap_position.x0 + shift,
                heatmap_position.y0,
                heatmap_position.width,
                heatmap_position.height,
            ]
        )
        label_left = get_y_tick_label_left()
        metric_group_right = label_left - strip_gap_x
        metric_group_left = metric_group_right - strip_width

    heatmap_position = ax_heatmap.get_position()
    ax_metric_group = None
    if show_metric_groups:
        ax_metric_group = fig.add_axes(
            [
                metric_group_left,
                heatmap_position.y0,
                strip_width,
                heatmap_position.height,
            ]
        )
        ax_metric_group.imshow(
            np.arange(n_rows).reshape(-1, 1),
            aspect="auto",
            cmap=metric_group_cmap,
        )
        ax_metric_group.set_ylim(ax_heatmap.get_ylim())
        ax_metric_group.set_xticks([])
        ax_metric_group.set_yticks([])
        ax_metric_group.tick_params(axis="y", left=False, labelleft=False, length=0)
        ax_metric_group.tick_params(axis="x", bottom=False, labelbottom=False)
        ax_metric_group.set_ylabel(
            text_labels["metric_group_strip_title"],
            fontsize=10,
            rotation=90,
            labelpad=8,
            va="center",
            fontweight="bold",
        )
        for spine in ax_metric_group.spines.values():
            spine.set_visible(False)

    def get_top_tick_label_edge() -> float:
        fig.canvas.draw()
        renderer = fig.canvas.get_renderer()
        tick_label_tops = [
            label.get_window_extent(renderer=renderer).transformed(
                fig.transFigure.inverted()
            ).y1
            for label in ax_heatmap.get_xticklabels()
            if label.get_visible()
        ]
        return max(tick_label_tops, default=ax_heatmap.get_position().y1)

    label_top = get_top_tick_label_edge()
    strip_bottom = label_top + strip_gap_y
    max_strip_bottom = 1 - strip_height - 0.45 / fig_height
    if strip_bottom > max_strip_bottom:
        heatmap_position = ax_heatmap.get_position()
        overflow = strip_bottom - max_strip_bottom
        new_height = max(0.25, heatmap_position.height - overflow)
        ax_heatmap.set_position(
            [
                heatmap_position.x0,
                heatmap_position.y0,
                heatmap_position.width,
                new_height,
            ]
        )
        if ax_metric_group is not None:
            metric_group_position = ax_metric_group.get_position()
            ax_metric_group.set_position(
                [
                    metric_group_position.x0,
                    metric_group_position.y0,
                    metric_group_position.width,
                    new_height,
                ]
            )
        label_top = get_top_tick_label_edge()
        strip_bottom = label_top + strip_gap_y

    heatmap_position = ax_heatmap.get_position()
    ax_class = fig.add_axes(
        [
            heatmap_position.x0,
            min(strip_bottom, max_strip_bottom),
            heatmap_position.width,
            strip_height,
        ]
    )
    ax_class.set_xlim(ax_heatmap.get_xlim())
    ax_class.set_ylim(-0.5, 0.5)
    ax_class.set_facecolor("white")
    for col_idx, x_position in enumerate(x_positions):
        ax_class.add_patch(
            mpatches.Rectangle(
                (x_position - 0.5, -0.5),
                1.0,
                1.0,
                facecolor=group_colors[col_idx],
                edgecolor="none",
                zorder=1,
            )
        )
    ax_class.set_xticks([])
    ax_class.set_yticks([])
    ax_class.tick_params(
        left=False,
        bottom=False,
        labelleft=False,
        labelbottom=False,
    )

    for start, end, group_label in compound_segments:
        group_color = compound_type_color(group_label)
        left = x_positions[start] - 0.5
        right = x_positions[end - 1] + 0.5
        center = (left + right) / 2
        ax_class.text(
            center,
            1.85,
            group_label,
            transform=ax_class.get_xaxis_transform(),
            ha="center",
            va="bottom",
            fontsize=9.4,
            fontweight="bold",
            color=group_color,
            clip_on=False,
        )
        ax_heatmap.hlines(
            y=-0.5,
            xmin=left,
            xmax=right,
            colors=group_color,
            linewidth=2.4,
            clip_on=False,
            zorder=5,
        )

    for spine in ax_class.spines.values():
        spine.set_visible(False)

    heatmap_position = ax_heatmap.get_position()
    ax_legend.set_position(
        [
            heatmap_position.x0,
            legend_bottom_space / fig_height,
            heatmap_position.width,
            legend_space / fig_height,
        ]
    )
    ax_legend.axis("off")

    fc_handles = [
        mpatches.Patch(facecolor=color_steps[0], edgecolor="black", linewidth=0.5, label=text_labels["bin_strong_down"]),
        mpatches.Patch(facecolor=color_steps[1], edgecolor="black", linewidth=0.5, label=text_labels["bin_mild_down"]),
        mpatches.Patch(facecolor=color_steps[2], edgecolor="black", linewidth=0.5, label=text_labels["bin_neutral"]),
        mpatches.Patch(facecolor=color_steps[3], edgecolor="black", linewidth=0.5, label=text_labels["bin_mild_up"]),
        mpatches.Patch(facecolor=color_steps[4], edgecolor="black", linewidth=0.5, label=text_labels["bin_strong_up"]),
    ]
    size_handles = [
        Line2D(
            [0],
            [0],
            marker="o",
            linestyle="",
            color="black",
            markerfacecolor="#bfbfbf",
            markeredgecolor="black",
            markersize=np.sqrt(dot_sizes["large"]),
            label=f"|Log2FC| >= {size_cutoff:g}",
        ),
        Line2D(
            [0],
            [0],
            marker="o",
            linestyle="",
            color="black",
            markerfacecolor="#bfbfbf",
            markeredgecolor="black",
            markersize=np.sqrt(dot_sizes["medium"]),
            label=f"0.5 <= |Log2FC| < {size_cutoff:g}",
        ),
        Line2D(
            [0],
            [0],
            marker="o",
            linestyle="",
            color="black",
            markerfacecolor="#bfbfbf",
            markeredgecolor="black",
            markersize=np.sqrt(dot_sizes["small"]),
            label="|Log2FC| < 0.5",
        ),
    ]
    class_handles = [
        mpatches.Patch(facecolor=color, edgecolor="black", linewidth=0.5, label=group)
        for group, color in group_to_color.items()
    ]
    metric_group_handles = [
        mpatches.Patch(facecolor=color, edgecolor="black", linewidth=0.5, label=group)
        for group, color in metric_group_to_color.items()
    ]

    def place_legend(
        handles: list[mpatches.Patch] | list[Line2D],
        title: str,
        anchor_x: float,
        anchor_y: float,
        columns: int = 1,
    ) -> plt.Legend:
        legend = ax_legend.legend(
            handles=handles,
            loc="upper left",
            bbox_to_anchor=(anchor_x, anchor_y),
            borderaxespad=0,
            borderpad=0,
            title=title,
            frameon=False,
            fontsize=8.2,
            title_fontsize=9.2,
            ncol=columns,
            columnspacing=0.9,
            handlelength=1.5,
            handletextpad=0.5,
            labelspacing=1.05 if title == text_labels.get("size_legend_title", "Circle size") else 0.35,
            handleheight=1.8 if title == text_labels.get("size_legend_title", "Circle size") else 0.7,
        )
        legend._legend_box.align = "left"
        legend.get_title().set_ha("left")
        return legend

    fc_legend = place_legend(fc_handles, text_labels["colorbar_title"], 0.0, 1.0)
    ax_legend.add_artist(fc_legend)
    size_legend = place_legend(
        size_handles,
        text_labels.get("size_legend_title", "Circle size"),
        0.22,
        1.0,
    )
    ax_legend.add_artist(size_legend)
    class_columns = 2 if len(class_handles) > 6 else 1
    class_legend = place_legend(
        class_handles,
        text_labels["legend_title"],
        0.44,
        1.0,
        columns=class_columns,
    )
    if show_metric_groups:
        ax_legend.add_artist(class_legend)
        metric_group_legend = place_legend(
            metric_group_handles,
            text_labels["metric_group_legend_title"],
            0.0,
            0.42,
        )

    return fig


def export_figure_png_bytes(fig: plt.Figure, dpi: int = 1000) -> bytes:
    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="png",
        dpi=dpi,
        bbox_inches="tight",
        facecolor="white",
    )
    buffer.seek(0)
    return buffer.getvalue()


def export_figure_tiff_bytes(fig: plt.Figure, dpi: int = 600) -> bytes:
    buffer = BytesIO()
    fig.savefig(
        buffer,
        format="tiff",
        dpi=dpi,
        bbox_inches="tight",
        facecolor="white",
        pil_kwargs={"compression": "tiff_lzw"},
    )
    buffer.seek(0)
    return buffer.getvalue()


def format_long_table(
    data: pd.DataFrame,
    metrics: list[str],
    metric_groups: dict[str, str] | None = None,
) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for _, row in data.iterrows():
        for metric in metrics:
            record = {
                "Class": row["group"],
                "Drug": row["drug"],
                "Metric": metric,
                "Log2FoldChange": row[f"{metric}__log2fc"],
                "p-value": row[f"{metric}__pvalue"],
            }
            if metric_groups is not None:
                record["Metric group"] = metric_groups.get(metric, "")
            records.append(record)
    return pd.DataFrame(records)


def format_compound_long_table(
    data: pd.DataFrame,
    metrics: list[str],
    metric_groups: dict[str, str] | None = None,
) -> pd.DataFrame:
    records: list[dict[str, object]] = []
    for _, row in data.iterrows():
        for metric in metrics:
            record = {
                "Compound type": row["compound_type"],
                "Class": row["group"],
                "Drug": row["drug"],
                "Metric": metric,
                "Log2FoldChange": row[f"{metric}__log2fc"],
                "p-value": row[f"{metric}__pvalue"],
            }
            if metric_groups is not None:
                record["Metric group"] = metric_groups.get(metric, "")
            records.append(record)
    return pd.DataFrame(records)


def render_standard_heatmap_tab() -> None:
    st.subheader("Классическая heatmap")
    st.caption(
        "Загрузите `.xlsx` с заданной структурой: колонка A — класс препарата, "
        "колонка B — препарат, дальше пары `Log2Fold Change` и `p-value` для каждой метрики."
    )
    uploaded_file = st.file_uploader("Загрузите `.xlsx`", type=["xlsx"], key="xlsx_uploader")
    file_bytes: bytes | None = uploaded_file.getvalue() if uploaded_file is not None else None

    if not file_bytes:
        return

    try:
        parsed = parse_workbook(file_bytes)
    except Exception as exc:
        st.error(
            "Не удалось прочитать файл. Проверьте, что это `.xlsx` нужной структуры, "
            "и попробуйте загрузить его ещё раз."
        )
        st.exception(exc)
        return

    data = parsed.data
    metrics = parsed.metrics
    metric_groups_by_metric = parsed.metric_groups

    if data.empty or not metrics:
        st.error("Не удалось извлечь таблицу из Excel.")
        return

    groups = data["group"].dropna().unique().tolist()

    with st.sidebar:
        st.header("Настройки")
        selected_groups = st.multiselect(
            "Классы препаратов",
            options=groups,
            default=groups,
        )
        selected_metrics = st.multiselect(
            "Метрики",
            options=metrics,
            default=metrics,
        )
        pvalue_threshold = st.selectbox(
            "Порог p-value",
            options=[0.01, 0.05],
            index=1,
            format_func=lambda value: f"{value:.2f}",
        )
        significant_only = st.checkbox(
            "Показывать только строки, где есть хотя бы одна значимая ячейка",
            value=False,
        )
        annotate = st.checkbox("Подписывать значения в ячейках", value=True)
        heatmap_orientation = st.selectbox(
            "Ориентация",
            options=["groups_top", "metrics_top"],
            index=0,
            format_func=lambda value: {
                "groups_top": "Препараты сверху, метрики слева",
                "metrics_top": "Метрики сверху, препараты слева",
            }[value],
        )
        fc_abs_cutoff = st.selectbox(
            "Порог |Log2FC|",
            options=[1.0, 0.58],
            index=0,
            format_func=lambda value: (
                f"-{value:g} / +{value:g}"
            ),
        )
        low_cutoff = -fc_abs_cutoff
        high_cutoff = fc_abs_cutoff

        st.divider()
        st.subheader("Редактирование подписей")
        default_text_labels = build_default_text_labels(fc_abs_cutoff, pvalue_threshold)
        sync_default_text_labels(default_text_labels)

        if st.button("Сбросить подписи по умолчанию", width="stretch"):
            for key, value in default_text_labels.items():
                st.session_state[f"text_{key}"] = value
            st.session_state["text_default_tracker"] = default_text_labels.copy()

        text_labels = {
            key: editable_text_input(label, f"text_{key}", default_text_labels[key])
            for key, label in [
                ("heatmap_title", "Заголовок графика"),
                ("class_strip_title", "Заголовок цветовой полосы классов"),
                ("legend_title", "Заголовок легенды"),
                ("metric_group_strip_title", "Заголовок цветовой полосы групп параметров"),
                ("metric_group_legend_title", "Заголовок легенды групп параметров"),
                ("colorbar_title", "Заголовок легенды fold change"),
                ("pvalue_note", "Подпись для незначимых ячеек"),
                ("bin_strong_down", "Подпись сильного снижения"),
                ("bin_mild_down", "Подпись умеренного снижения"),
                ("bin_neutral", "Подпись нейтральной зоны"),
                ("bin_mild_up", "Подпись умеренного повышения"),
                ("bin_strong_up", "Подпись сильного повышения"),
            ]
        }

    if not selected_groups:
        st.warning("Выберите хотя бы один класс препаратов.")
        return
    if not selected_metrics:
        st.warning("Выберите хотя бы одну метрику.")
        return

    filtered_meta, value_frame, pvalue_frame = build_heatmap_frame(
        data=data,
        selected_groups=selected_groups,
        selected_metrics=selected_metrics,
        pvalue_threshold=pvalue_threshold,
        significant_only=significant_only,
    )

    if value_frame.empty:
        st.warning("После фильтрации не осталось данных для отображения.")
        return

    metric_group_names = [
        group
        for group in dict.fromkeys(
            metric_groups_by_metric.get(metric, "") for metric in selected_metrics
        )
        if group
    ]
    metric_group_label_map: dict[str, str] = {}

    with st.expander("Редактирование текстов на графике", expanded=False):
        metric_label_map = editable_mapping_editor(
            items=selected_metrics,
            state_key="metric_label_map",
            source_label="Исходная метрика",
            target_label="Подпись на графике",
        )
        if parsed.has_metric_groups:
            metric_group_label_map = editable_mapping_editor(
                items=metric_group_names,
                state_key="metric_group_label_map",
                source_label="Исходная группа параметров",
                target_label="Подпись на графике",
            )
        group_label_map = editable_mapping_editor(
            items=selected_groups,
            state_key="group_label_map",
            source_label="Исходный класс",
            target_label="Подпись на графике",
        )
        drug_label_map = editable_mapping_editor(
            items=filtered_meta["drug"].tolist(),
            state_key="drug_label_map",
            source_label="Исходный препарат",
            target_label="Подпись на графике",
        )

    metric_labels = [metric_label_map.get(metric, metric) for metric in selected_metrics]
    metric_group_labels = [
        metric_group_label_map.get(
            metric_groups_by_metric.get(metric, ""),
            metric_groups_by_metric.get(metric, ""),
        )
        for metric in selected_metrics
    ]
    show_metric_groups = parsed.has_metric_groups and any(metric_group_labels)
    group_labels = [group_label_map.get(group, group) for group in filtered_meta["group"]]
    drug_labels = [drug_label_map.get(drug, drug) for drug in filtered_meta["drug"]]
    visible_groups = filtered_meta["group"].dropna().unique().tolist()

    with st.sidebar:
        st.divider()
        st.subheader("Цвета")
        with st.expander("Ячейки heatmap", expanded=False):
            heatmap_colors = [
                st.color_picker(
                    text_labels["bin_strong_down"],
                    value=DEFAULT_HEATMAP_COLORS["strong_down"],
                    key=f"color_heatmap_strong_down_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_mild_down"],
                    value=DEFAULT_HEATMAP_COLORS["mild_down"],
                    key=f"color_heatmap_mild_down_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_neutral"],
                    value=DEFAULT_HEATMAP_COLORS["neutral"],
                    key=f"color_heatmap_neutral_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_mild_up"],
                    value=DEFAULT_HEATMAP_COLORS["mild_up"],
                    key=f"color_heatmap_mild_up_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_strong_up"],
                    value=DEFAULT_HEATMAP_COLORS["strong_up"],
                    key=f"color_heatmap_strong_up_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
            ]
            pvalue_color = st.color_picker(
                text_labels["pvalue_note"],
                value=DEFAULT_HEATMAP_COLORS["pvalue"],
                key=f"color_heatmap_pvalue_{HEATMAP_COLOR_WIDGET_VERSION}",
            )

        with st.expander("Шкала Class", expanded=False):
            class_colors_by_group = editable_color_map(
                visible_groups,
                build_ordered_color_map(
                    visible_groups,
                    DEFAULT_CLASS_COLORS,
                    "tab20",
                ),
                f"color_class_{CLASS_COLOR_WIDGET_VERSION}",
                label_map=group_label_map,
            )

        metric_group_colors_by_group: dict[str, str] = {}
        if show_metric_groups:
            with st.expander("Шкала Parameter group", expanded=False):
                metric_group_colors_by_group = editable_color_map(
                    metric_group_names,
                    build_ordered_color_map(
                        metric_group_names,
                        DEFAULT_METRIC_GROUP_COLORS,
                        "Set2",
                    ),
                    f"color_metric_group_{METRIC_GROUP_COLOR_WIDGET_VERSION}",
                    label_map=metric_group_label_map,
                )

    class_color_map = {
        group_label_map.get(group, group): color
        for group, color in class_colors_by_group.items()
    }
    metric_group_color_map = {
        metric_group_label_map.get(group, group): color
        for group, color in metric_group_colors_by_group.items()
    }

    st.subheader("Heatmap")
    fig = plot_heatmap(
        values=value_frame,
        pvalues=pvalue_frame,
        pvalue_threshold=pvalue_threshold,
        low_cutoff=low_cutoff,
        high_cutoff=high_cutoff,
        annotate=annotate,
        text_labels=text_labels,
        metric_labels=metric_labels,
        metric_group_labels=metric_group_labels,
        drug_labels=drug_labels,
        group_labels=group_labels,
        show_metric_groups=show_metric_groups,
        heatmap_colors=heatmap_colors,
        pvalue_color=pvalue_color,
        class_color_map=class_color_map,
        metric_group_color_map=metric_group_color_map,
        orientation=heatmap_orientation,
    )
    st.pyplot(fig, width="stretch")

    export_png_col, export_tiff_col, _ = st.columns([1, 1, 3])
    with export_png_col:
        st.download_button(
            "Скачать PNG (1000 dpi)",
            data=export_figure_png_bytes(fig, dpi=1000),
            file_name="behavior_heatmap_1000dpi.png",
            mime="image/png",
            width="stretch",
        )
    with export_tiff_col:
        st.download_button(
            "Скачать TIFF (600 dpi)",
            data=export_figure_tiff_bytes(fig, dpi=600),
            file_name="behavior_heatmap_600dpi.tiff",
            mime="image/tiff",
            width="stretch",
        )

    summary_1, summary_2, summary_3 = st.columns(3)
    summary_1.metric("Препаратов", len(filtered_meta))
    summary_2.metric("Метрик", len(selected_metrics))
    significant_cells = int((pvalue_frame <= pvalue_threshold).fillna(False).sum().sum())
    summary_3.metric("Значимых ячеек", significant_cells)

    with st.expander("Таблица значений"):
        st.dataframe(
            format_long_table(
                filtered_meta,
                selected_metrics,
                metric_groups_by_metric if parsed.has_metric_groups else None,
            ),
            width="stretch",
            hide_index=True,
        )


def render_compound_dot_heatmap_tab() -> None:
    st.subheader("Кружковая матрица")
    st.caption(
        "Загрузите `.xlsx` новой структуры: колонка A — тип соединения, "
        "колонка B — класс препарата, колонка C — препарат, дальше пары "
        "`Log2Fold Change` и `p-value` для каждой метрики."
    )
    uploaded_file = st.file_uploader(
        "Загрузите `.xlsx` для кружковой матрицы",
        type=["xlsx"],
        key="compound_xlsx_uploader",
    )
    file_bytes: bytes | None = uploaded_file.getvalue() if uploaded_file is not None else None

    if not file_bytes:
        return

    try:
        parsed = parse_compound_workbook(file_bytes)
    except Exception as exc:
        st.error(
            "Не удалось прочитать файл новой структуры. Проверьте, что колонка A — тип "
            "соединения, B — класс, C — препарат, а метрики идут парами Log2Fold Change / p-value."
        )
        st.exception(exc)
        return

    data = parsed.data
    metrics = parsed.metrics
    metric_groups_by_metric = parsed.metric_groups

    if data.empty or not metrics:
        st.error("Не удалось извлечь таблицу из Excel.")
        return

    compound_types = data["compound_type"].dropna().unique().tolist()
    groups = data["group"].dropna().unique().tolist()

    with st.sidebar:
        st.divider()
        st.header("Кружковая матрица")
        selected_compound_types = st.multiselect(
            "Типы соединений",
            options=compound_types,
            default=compound_types,
            key="dot_selected_compound_types",
        )
        selected_groups = st.multiselect(
            "Классы препаратов",
            options=groups,
            default=groups,
            key="dot_selected_groups",
        )
        selected_metrics = st.multiselect(
            "Метрики",
            options=metrics,
            default=metrics,
            key="dot_selected_metrics",
        )
        pvalue_threshold = st.selectbox(
            "Порог p-value",
            options=[0.01, 0.05],
            index=1,
            format_func=lambda value: f"{value:.2f}",
            key="dot_pvalue_threshold",
        )
        significant_only = st.checkbox(
            "Только строки с хотя бы одной значимой ячейкой",
            value=False,
            key="dot_significant_only",
        )
        fc_abs_cutoff = st.selectbox(
            "Порог |Log2FC|",
            options=[1.0, 0.58],
            index=0,
            format_func=lambda value: f"-{value:g} / +{value:g}",
            key="dot_fc_abs_cutoff",
        )

        st.divider()
        st.subheader("Подписи кружковой матрицы")
        default_text_labels = build_default_text_labels(fc_abs_cutoff, pvalue_threshold)
        default_text_labels["size_legend_title"] = "Circle size"
        sync_default_text_labels(
            default_text_labels,
            state_prefix="dot_text_",
            tracker_key="dot_text_default_tracker",
        )

        if st.button("Сбросить подписи кружковой матрицы", width="stretch"):
            for key, value in default_text_labels.items():
                st.session_state[f"dot_text_{key}"] = value
            st.session_state["dot_text_default_tracker"] = default_text_labels.copy()

        text_labels = {
            key: editable_text_input(label, f"dot_text_{key}", default_text_labels[key])
            for key, label in [
                ("heatmap_title", "Заголовок графика"),
                ("class_strip_title", "Заголовок шкалы классов"),
                ("legend_title", "Заголовок легенды классов"),
                ("metric_group_strip_title", "Заголовок шкалы functional groups"),
                ("metric_group_legend_title", "Заголовок легенды functional groups"),
                ("colorbar_title", "Заголовок легенды fold change"),
                ("size_legend_title", "Заголовок легенды размера кружков"),
                ("bin_strong_down", "Подпись сильного снижения"),
                ("bin_mild_down", "Подпись умеренного снижения"),
                ("bin_neutral", "Подпись нейтральной зоны"),
                ("bin_mild_up", "Подпись умеренного повышения"),
                ("bin_strong_up", "Подпись сильного повышения"),
            ]
        }

    low_cutoff = -fc_abs_cutoff
    high_cutoff = fc_abs_cutoff

    if not selected_compound_types:
        st.warning("Выберите хотя бы один тип соединений.")
        return
    if not selected_groups:
        st.warning("Выберите хотя бы один класс препаратов.")
        return
    if not selected_metrics:
        st.warning("Выберите хотя бы одну метрику.")
        return

    filtered_meta, value_frame, pvalue_frame = build_compound_heatmap_frame(
        data=data,
        selected_compound_types=selected_compound_types,
        selected_groups=selected_groups,
        selected_metrics=selected_metrics,
        pvalue_threshold=pvalue_threshold,
        significant_only=significant_only,
    )

    if value_frame.empty:
        st.warning("После фильтрации не осталось данных для отображения.")
        return

    metric_group_names = [
        group
        for group in dict.fromkeys(
            metric_groups_by_metric.get(metric, "") for metric in selected_metrics
        )
        if group
    ]
    show_metric_groups = parsed.has_metric_groups and bool(metric_group_names)

    with st.sidebar:
        with st.expander("Редактирование текстов кружковой матрицы", expanded=False):
            metric_label_map = editable_mapping_editor(
                items=selected_metrics,
                state_key="dot_metric_label_map",
                source_label="Исходная метрика",
                target_label="Подпись на графике",
            )
            if parsed.has_metric_groups:
                metric_group_label_map = editable_mapping_editor(
                    items=metric_group_names,
                    state_key="dot_metric_group_label_map",
                    source_label="Исходная группа параметров",
                    target_label="Подпись на графике",
                )
            else:
                metric_group_label_map = {}
            group_label_map = editable_mapping_editor(
                items=selected_groups,
                state_key="dot_group_label_map",
                source_label="Исходный класс",
                target_label="Подпись на графике",
            )
            compound_type_label_map = editable_mapping_editor(
                items=selected_compound_types,
                state_key="dot_compound_type_label_map",
                source_label="Исходный тип",
                target_label="Подпись на графике",
            )
            drug_label_map = editable_mapping_editor(
                items=filtered_meta["drug"].tolist(),
                state_key="dot_drug_label_map",
                source_label="Исходный препарат",
                target_label="Подпись на графике",
            )

    metric_labels = [metric_label_map.get(metric, metric) for metric in selected_metrics]
    metric_group_labels = [
        metric_group_label_map.get(
            metric_groups_by_metric.get(metric, ""),
            metric_groups_by_metric.get(metric, ""),
        )
        for metric in selected_metrics
    ]
    show_metric_groups = parsed.has_metric_groups and any(metric_group_labels)
    drug_labels = [
        drug_label_map.get(drug, drug)
        for drug in filtered_meta["drug"].tolist()
    ]
    group_labels = [
        group_label_map.get(group, group)
        for group in filtered_meta["group"].tolist()
    ]
    compound_type_labels = [
        compound_type_label_map.get(compound_type, compound_type)
        for compound_type in filtered_meta["compound_type"].tolist()
    ]
    visible_groups = filtered_meta["group"].dropna().unique().tolist()
    visible_compound_types = filtered_meta["compound_type"].dropna().unique().tolist()

    with st.sidebar:
        st.divider()
        st.subheader("Цвета кружковой матрицы")
        with st.expander("Кружки", expanded=False):
            heatmap_colors = [
                st.color_picker(
                    text_labels["bin_strong_down"],
                    value=DEFAULT_HEATMAP_COLORS["strong_down"],
                    key=f"dot_color_heatmap_strong_down_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_mild_down"],
                    value=DEFAULT_HEATMAP_COLORS["mild_down"],
                    key=f"dot_color_heatmap_mild_down_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_neutral"],
                    value=DEFAULT_HEATMAP_COLORS["neutral"],
                    key=f"dot_color_heatmap_neutral_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_mild_up"],
                    value=DEFAULT_HEATMAP_COLORS["mild_up"],
                    key=f"dot_color_heatmap_mild_up_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
                st.color_picker(
                    text_labels["bin_strong_up"],
                    value=DEFAULT_HEATMAP_COLORS["strong_up"],
                    key=f"dot_color_heatmap_strong_up_{HEATMAP_COLOR_WIDGET_VERSION}",
                ),
            ]
            pvalue_color = DEFAULT_HEATMAP_COLORS["pvalue"]

        with st.expander("Шкала Class", expanded=False):
            class_colors_by_group = editable_color_map(
                visible_groups,
                build_ordered_color_map(
                    visible_groups,
                    DEFAULT_CLASS_COLORS,
                    "tab20",
                ),
                f"dot_color_class_{CLASS_COLOR_WIDGET_VERSION}",
                label_map=group_label_map,
            )

        metric_group_colors_by_group: dict[str, str] = {}
        if show_metric_groups:
            with st.expander("Шкала Parameter group", expanded=False):
                metric_group_colors_by_group = editable_color_map(
                    metric_group_names,
                    build_ordered_color_map(
                        metric_group_names,
                        DEFAULT_METRIC_GROUP_COLORS,
                        "Set2",
                    ),
                    f"dot_color_metric_group_{METRIC_GROUP_COLOR_WIDGET_VERSION}",
                    label_map=metric_group_label_map,
                )

    class_color_map = {
        group_label_map.get(group, group): color
        for group, color in class_colors_by_group.items()
    }
    metric_group_color_map = {
        metric_group_label_map.get(group, group): color
        for group, color in metric_group_colors_by_group.items()
    }

    fig = plot_compound_dot_heatmap(
        values=value_frame,
        pvalues=pvalue_frame,
        pvalue_threshold=pvalue_threshold,
        low_cutoff=low_cutoff,
        high_cutoff=high_cutoff,
        text_labels=text_labels,
        metric_labels=metric_labels,
        metric_group_labels=metric_group_labels,
        drug_labels=drug_labels,
        group_labels=group_labels,
        compound_type_labels=compound_type_labels,
        show_metric_groups=show_metric_groups,
        heatmap_colors=heatmap_colors,
        pvalue_color=pvalue_color,
        class_color_map=class_color_map,
        metric_group_color_map=metric_group_color_map,
    )
    st.pyplot(fig, width="stretch")

    export_png_col, export_tiff_col, _ = st.columns([1, 1, 3])
    with export_png_col:
        st.download_button(
            "Скачать PNG (1000 dpi)",
            data=export_figure_png_bytes(fig, dpi=1000),
            file_name="compound_dot_heatmap_1000dpi.png",
            mime="image/png",
            width="stretch",
            key="dot_download_png",
        )
    with export_tiff_col:
        st.download_button(
            "Скачать TIFF (600 dpi)",
            data=export_figure_tiff_bytes(fig, dpi=600),
            file_name="compound_dot_heatmap_600dpi.tiff",
            mime="image/tiff",
            width="stretch",
            key="dot_download_tiff",
        )

    summary_1, summary_2, summary_3, summary_4 = st.columns(4)
    summary_1.metric("Типов", len(visible_compound_types))
    summary_2.metric("Препаратов", len(filtered_meta))
    summary_3.metric("Метрик", len(selected_metrics))
    significant_cells = int((pvalue_frame <= pvalue_threshold).fillna(False).sum().sum())
    summary_4.metric("Значимых ячеек", significant_cells)

    with st.expander("Таблица значений"):
        st.dataframe(
            format_compound_long_table(
                filtered_meta,
                selected_metrics,
                metric_groups_by_metric if parsed.has_metric_groups else None,
            ),
            width="stretch",
            hide_index=True,
        )


def main() -> None:
    st.set_page_config(page_title="Behavior Heatmap", layout="wide")
    st.title("Heatmap по данным из Excel")
    standard_tab, compound_tab = st.tabs(
        ["Классическая heatmap", "Кружковая матрица"]
    )
    with standard_tab:
        render_standard_heatmap_tab()
    with compound_tab:
        render_compound_dot_heatmap_tab()


if __name__ == "__main__":
    main()
